import tkinter as Tk
from tkinter import ttk, messagebox, filedialog
import pymysql
from pymysql.cursors import DictCursor
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from tkcalendar import DateEntry
import threading
import time

class Database:
    connection_data = dict(
        host='localhost',
        user='root',
        password='',
        database='center',
        cursorclass=DictCursor
    )
    
    def __init__(self):
        self.db = pymysql.connect(**self.connection_data)
        self.cursor = self.db.cursor()
    
    def get_table_columns(self, table):
        self.cursor.execute(f"SHOW COLUMNS FROM {table}")
        return [column['Field'] for column in self.cursor.fetchall()]
    
    def get_table_data(self, table):
        try:
            if table == 'activation':
                query = '''SELECT a.id, c.full_name, c.phone, a.activation_datetime, a.completed
                        FROM activation a
                        JOIN contact c ON a.contact_id = c.id'''
            elif table == 'repair':
                query = '''SELECT r.id, c.full_name, c.phone, t.serial_number, 
                        r.repair_datetime, u.username 
                        FROM repair r
                        JOIN contact c ON r.contact_id = c.id
                        JOIN tachograph t ON r.tachograph_id = t.id
                        JOIN users u ON r.user_id = u.id'''
            elif table == 'calibration':
                query = '''SELECT cal.id, t.serial_number, cal.calibration_date,
                        cal.next_calibration_date, u.username, cal.completed 
                        FROM calibration cal
                        JOIN tachograph t ON cal.tachograph_id = t.id
                        JOIN users u ON cal.user_id = u.id'''
            else:
                query = f'SELECT * FROM {table}'
            
            self.cursor.execute(query)
            return self.cursor.fetchall()
        except pymysql.Error as e:
            messagebox.showerror("Ошибка базы данных", str(e))
            return []

    def add_record(self, table, data):
        try:
            columns = ', '.join(data.keys())
            placeholders = ', '.join(['%s'] * len(data))
            query = f"INSERT INTO {table} ({columns}) VALUES ({placeholders})"
            self.cursor.execute(query, list(data.values()))
            self.db.commit()
            return True
        except pymysql.Error as e:
            self.db.rollback()
            messagebox.showerror("Ошибка базы данных", str(e))
            return False

    def delete_record(self, table, record_id):
        try:
            query = f"DELETE FROM {table} WHERE id = %s"
            self.cursor.execute(query, (record_id,))
            self.db.commit()
            return True
        except pymysql.Error as e:
            self.db.rollback()
            messagebox.showerror("Ошибка базы данных", str(e))
            return False

    def update_record(self, table, record_id, data):
        try:
            set_clause = ', '.join([f"{key} = %s" for key in data.keys()])
            query = f"UPDATE {table} SET {set_clause} WHERE id = %s"
            self.cursor.execute(query, list(data.values()) + [record_id])
            self.db.commit()
            return True
        except pymysql.Error as e:
            self.db.rollback()
            messagebox.showerror("Ошибка базы данных", str(e))
            return False

    def get_record(self, table, record_id):
        try:
            query = f"SELECT * FROM {table} WHERE id = %s"
            self.cursor.execute(query, (record_id,))
            return self.cursor.fetchone()
        except pymysql.Error as e:
            messagebox.showerror("Ошибка базы данных", str(e))
            return None
    
    def get_calibration_details(self, calibration_id):
        try:
            query = '''SELECT 
                cal.*, 
                t.serial_number, 
                u.username,
                c.full_name,
                c.phone,
                v.brand,
                v.model,
                v.gos_number
            FROM calibration cal
            JOIN tachograph t ON cal.tachograph_id = t.id
            JOIN users u ON cal.user_id = u.id
            JOIN vehicle v ON t.vehicle_id = v.id 
            JOIN contact c ON t.contact_id = c.id
            WHERE cal.id = %s'''
            self.cursor.execute(query, (calibration_id,))
            return self.cursor.fetchone()
        except pymysql.Error as e:
            messagebox.showerror("Ошибка базы данных", str(e))
            return None

    def get_repair_details(self, repair_id):
        try:
            query = '''SELECT 
                r.*,
                c.full_name,
                c.phone,
                t.serial_number,
                u.username,
                v.brand,
                v.model,
                v.gos_number
                FROM repair r
                JOIN contact c ON r.contact_id = c.id
                JOIN tachograph t ON r.tachograph_id = t.id
                JOIN users u ON r.user_id = u.id
                JOIN vehicle v ON t.Vehicle_id = v.id
                WHERE r.id = %s'''
            self.cursor.execute(query, (repair_id,))
            return self.cursor.fetchone()
        except pymysql.Error as e:
            messagebox.showerror("Ошибка базы данных", str(e))
            return None
        
    def mark_procedure_completed(self, table, record_id):
        try:
            query = f"UPDATE {table} SET completed = TRUE WHERE id = %s"
            self.cursor.execute(query, (record_id,))
            self.db.commit()
            return True
        except pymysql.Error as e:
            self.db.rollback()
            messagebox.showerror("Ошибка базы данных", str(e))
            return False
    
    def delete_passport_by_contact(self, contact_id):
            try:
                query = "DELETE FROM passport WHERE contact_id = %s"
                self.cursor.execute(query, (contact_id,))
                self.db.commit()
                return True
            except pymysql.Error as e:
                self.db.rollback()
                print(f"Ошибка удаления паспорта: {str(e)}")
                return False

    def get_notifications(self):
        try:
            notifications = []
            
            if not self.db.open:
                self.db.ping(reconnect=True)
                
            self.cursor.execute('''
                SELECT 
                    'calibration' as type, 
                    id, 
                    serial_number, 
                    DATE(next_calibration_date) as next_calibration_date  
                FROM (
                    SELECT 
                        c.id, 
                        t.serial_number, 
                        c.next_calibration_date
                    FROM calibration c
                    JOIN tachograph t ON c.tachograph_id = t.id
                    WHERE c.completed = FALSE
                ) as subquery
                WHERE next_calibration_date BETWEEN CURDATE() AND DATE_ADD(CURDATE(), INTERVAL 14 DAY)
            ''')
            calibration_notifications = self.cursor.fetchall()
            notifications.extend(calibration_notifications)
            
            return notifications
        except Exception as e:
            print(f"Ошибка получения уведомлений: {str(e)}")
            return []
        
    def close(self):
        try:
            if self.cursor:
                self.cursor.close()
            if self.db and self.db.open:
                self.db.close()
        except Exception as e:
            print(f"Ошибка при закрытии соединения: {str(e)}")

class LoginWindow(Tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("Авторизация")
        self.geometry("300x200")
        self.user_id = None
        self.account_type = None
        self.success = False
        
        Tk.Label(self, text="Логин:").pack(pady=5)
        self.username_entry = Tk.Entry(self)
        self.username_entry.pack(pady=5)
        
        Tk.Label(self, text="Пароль:").pack(pady=5)
        self.password_entry = Tk.Entry(self, show="*")
        self.password_entry.pack(pady=5)
        
        self.login_button = ttk.Button(self, text="Войти", command=self.authenticate)
        self.login_button.pack(pady=5)
        
        self.bind('<Return>', lambda event: self.authenticate())

    def authenticate(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        if not username or not password:
            messagebox.showwarning("Ошибка", "Введите логин и пароль")
            return
        
        db = Database()
        try:
            db.cursor.execute(
                "SELECT id, account_type FROM users WHERE username = %s AND password = %s",
                (username, password)
            )
            result = db.cursor.fetchone()
            if result:
                self.user_id = result['id']
                self.account_type = result['account_type']
                self.success = True
                self.destroy()
            else:
                messagebox.showerror("Ошибка", "Неверный логин или пароль")
        except pymysql.Error as e:
            messagebox.showerror("Ошибка базы данных", str(e))
            self.destroy()
            self.parent.destroy()
        finally:
            db.db.close()

class AddView(Tk.Toplevel):
    def __init__(self, parent, table, columns):
        super().__init__(parent.root)
        self.parent = parent
        self.table = table
        self.columns = columns
        self.title(f"Добавить в {table}")
        self.geometry("400x500")
        
        self.entries = {}
        for idx, column in enumerate(self.columns):
            label_text = self.parent.column_translations[table].get(column, column)
            Tk.Label(self, text=label_text).grid(row=idx, column=0, padx=5, pady=5)
            
            if column in ['date_start', 'date_end'] and table == 'mrp':
                entry = DateEntry(self, 
                                date_pattern='yyyy-mm-dd',
                                locale='ru_RU')
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[column] = entry
            elif column == 'account_type' and table == 'users':
                combo = ttk.Combobox(self, values=['admin', 'operator', 'master'])
                combo.current(0)
                combo.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[column] = combo
            elif column in ['date_issued', 'date_of_birth'] and table == 'passport':
                entry = DateEntry(self, 
                                date_pattern='yyyy-mm-dd',
                                locale='ru_RU')
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[column] = entry
            else:
                entry = Tk.Entry(self)
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[column] = entry

        btn_save = ttk.Button(self, text="Сохранить", command=self.save)
        btn_save.grid(row=len(self.columns)+1, columnspan=2, pady=10)

    def save(self):
        data = {}
        for col, entry in self.entries.items():
            if col in ['date_start', 'date_end', 'date_issued', 'date_of_birth'] and isinstance(entry, DateEntry):
                data[col] = entry.get_date().strftime('%Y-%m-%d')
            elif col == 'account_type':
                data[col] = entry.get()
            else:
                data[col] = entry.get()
        
        if all(data.values()):
            if self.parent.database.add_record(self.table, data):
                self.parent.update_tree()
                self.destroy()
        else:
            messagebox.showwarning("Ошибка", "Заполните все поля")

class EditView(Tk.Toplevel):
    def __init__(self, parent, table, record):
        super().__init__(parent.root)
        self.parent = parent
        self.table = table
        self.record = record
        self.title(f"Редактировать {table}")
        self.geometry("400x500")
        
        self.entries = {}
        translations = self.parent.column_translations.get(table, {})
        for idx, (col, val) in enumerate(record.items()):
            if col == 'id': 
                continue
            
            label_text = translations.get(col, col)
            Tk.Label(self, text=label_text).grid(row=idx, column=0, padx=5, pady=5)
            
            if col in ['date_start', 'date_end'] and table == 'mrp':
                entry = DateEntry(self, 
                                 date_pattern='yyyy-mm-dd',
                                 locale='ru_RU')
                entry.set_date(datetime.strptime(str(val), '%Y-%m-%d'))
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[col] = entry
            elif col == 'account_type' and table == 'users':
                combo = ttk.Combobox(self, values=['admin', 'operator', 'master'])
                combo.set(val)
                combo.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[col] = combo
            elif col in ['date_issued', 'date_of_birth'] and table == 'passport':
                entry = DateEntry(self, 
                                 date_pattern='yyyy-mm-dd',
                                 locale='ru_RU')
                if val:  
                    entry.set_date(datetime.strptime(str(val), '%Y-%m-%d'))
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[col] = entry
            else:
                entry = Tk.Entry(self)
                entry.insert(0, str(val))
                entry.grid(row=idx, column=1, padx=5, pady=5)
                self.entries[col] = entry
        
        btn_frame = Tk.Frame(self)
        btn_frame.grid(row=len(self.entries)+1, columnspan=2, pady=10)
        
        ttk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=Tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side=Tk.LEFT, padx=5)
    
    def save(self):
        data = {}
        for col, entry in self.entries.items():
            if col in ['date_start', 'date_end', 'date_issued', 'date_of_birth'] and isinstance(entry, DateEntry):
                data[col] = entry.get_date().strftime('%Y-%m-%d')
            elif col == 'account_type':
                data[col] = entry.get()
            else:
                data[col] = entry.get()
        
        if all(data.values()):
            if self.parent.database.add_record(self.table, data):
                self.parent.update_tree()
                self.destroy()
        else:
            messagebox.showwarning("Ошибка", "Заполните все поля")

class AddActivationWindow(Tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent.root)
        self.parent = parent
        self.title("Новая активация")
        self.geometry("400x300")

        Tk.Label(self, text="Клиент:").pack(pady=5)
        self.contact_combobox = ttk.Combobox(self)
        self.contact_combobox.pack(pady=5)
        
        Tk.Label(self, text="Дата:").pack(pady=5)
        self.date_entry = DateEntry(self, 
                                  date_pattern='yyyy-mm-dd',
                                  locale='ru_RU')
        self.date_entry.pack(pady=5)
        
        Tk.Label(self, text="Время (ЧЧ:ММ):").pack(pady=5)
        self.time_entry = ttk.Entry(self)
        self.time_entry.pack(pady=5)
        self.time_entry.insert(0, datetime.now().strftime("%H:%M"))

        btn_frame = Tk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=Tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side=Tk.LEFT, padx=5)

        self.load_contacts()

    def load_contacts(self):
        try:
            self.parent.database.cursor.execute("SELECT id, full_name FROM contact")
            contacts = [f"{row['id']} - {row['full_name']}" for row in self.parent.database.cursor.fetchall()]
            self.contact_combobox['values'] = contacts
            if contacts: 
                self.contact_combobox.current(0)
        except pymysql.Error as e:
            messagebox.showerror("Ошибка", str(e))

    def save(self):
        contact_id = self.contact_combobox.get().split(" - ")[0]
        selected_date = self.date_entry.get_date()
        time_str = self.time_entry.get()
        
        try:
            datetime.strptime(time_str, "%H:%M")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат времени! Используйте ЧЧ:ММ")
            return
            
        activation_datetime = f"{selected_date} {time_str}"
        
        if not contact_id or not activation_datetime:
            messagebox.showwarning("Ошибка", "Заполните все поля")
            return
        
        try:
            self.parent.database.add_record('activation', {
                'contact_id': contact_id,
                'activation_datetime': activation_datetime,
                'completed': False
            })
            self.parent.load_activation_data()
            self.destroy()
            messagebox.showinfo("Успех", "Активация добавлена")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

class AddRepairWindow(Tk.Toplevel):
    def __init__(self, parent, user_id):
        super().__init__(parent.root)
        self.parent = parent
        self.user_id = user_id
        self.title("Новый ремонт")
        self.geometry("500x400")

        Tk.Label(self, text="Клиент:").pack(pady=5)
        self.contact_combobox = ttk.Combobox(self)
        self.contact_combobox.pack(pady=5)
 
        Tk.Label(self, text="Тахограф:").pack(pady=5)
        self.tacho_combobox = ttk.Combobox(self)
        self.tacho_combobox.pack(pady=5)
 
        Tk.Label(self, text="Дата:").pack(pady=5)
        self.date_entry = DateEntry(self, 
                                  date_pattern='yyyy-mm-dd',
                                  locale='ru_RU')
        self.date_entry.pack(pady=5)
        
        Tk.Label(self, text="Время (ЧЧ:ММ):").pack(pady=5)
        self.time_entry = ttk.Entry(self)
        self.time_entry.pack(pady=5)
        self.time_entry.insert(0, datetime.now().strftime("%H:%M"))

        btn_frame = Tk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=Tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side=Tk.LEFT, padx=5)
        
        self.load_contacts()
        self.load_tachographs()

    def load_contacts(self):
        try:
            self.parent.database.cursor.execute("SELECT id, full_name FROM contact")
            contacts = [f"{row['id']} - {row['full_name']}" for row in self.parent.database.cursor.fetchall()]
            self.contact_combobox['values'] = contacts
            if contacts: self.contact_combobox.current(0)
        except pymysql.Error as e:
            messagebox.showerror("Ошибка", str(e))
    
    def load_tachographs(self):
        try:
            self.parent.database.cursor.execute("SELECT id, serial_number FROM tachograph")
            tachos = [f"{row['id']} - {row['serial_number']}" for row in self.parent.database.cursor.fetchall()]
            self.tacho_combobox['values'] = tachos
            if tachos: self.tacho_combobox.current(0)
        except pymysql.Error as e:
            messagebox.showerror("Ошибка", str(e))

    def save(self):
        contact_id = self.contact_combobox.get().split(" - ")[0]
        tacho_id = self.tacho_combobox.get().split(" - ")[0]
        selected_date = self.date_entry.get_date()
        time_str = self.time_entry.get()
        
        try:
            datetime.strptime(time_str, "%H:%M")
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат времени! Используйте ЧЧ:ММ")
            return
            
        repair_datetime = f"{selected_date} {time_str}"
        
        if not all([contact_id, tacho_id, repair_datetime]):
            messagebox.showwarning("Ошибка", "Заполните все поля")
            return
        
        try:
            self.parent.database.add_record('repair', {
                'contact_id': contact_id,
                'tachograph_id': tacho_id,
                'repair_datetime': repair_datetime,
                'user_id': self.user_id
            })
            self.parent.load_repair_data()
            self.destroy()
            messagebox.showinfo("Успех", "Ремонт добавлен")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

class AddCalibrationWindow(Tk.Toplevel):
    def __init__(self, parent, user_id):
        super().__init__(parent.root)
        self.parent = parent
        self.user_id = user_id
        self.title("Новая калибровка")
        self.geometry("400x300")
        
        Tk.Label(self, text="Тахограф:").pack(pady=5)
        self.tacho_combobox = ttk.Combobox(self)
        self.tacho_combobox.pack(pady=5)
  
        Tk.Label(self, text="Дата калибровки:").pack(pady=5)
        self.calibration_date = DateEntry(self, 
                                        date_pattern='yyyy-mm-dd',
                                        locale='ru_RU')
        self.calibration_date.pack(pady=5)
        
        Tk.Label(self, text="Следующая калибровка:").pack(pady=5)
        self.next_calibration = DateEntry(self, 
                                        date_pattern='yyyy-mm-dd',
                                        locale='ru_RU')
        self.next_calibration.pack(pady=5)

        btn_frame = Tk.Frame(self)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=Tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Отмена", command=self.destroy).pack(side=Tk.LEFT, padx=5)
        
        self.load_tachographs()

    def load_tachographs(self):
        try:
            self.parent.database.cursor.execute("SELECT id, serial_number FROM tachograph")
            tachos = [f"{row['id']} - {row['serial_number']}" for row in self.parent.database.cursor.fetchall()]
            self.tacho_combobox['values'] = tachos
            if tachos: 
                self.tacho_combobox.current(0)
        except pymysql.Error as e:
            messagebox.showerror("Ошибка", str(e))
            self.destroy()


    def save(self):
        try:
            tacho_id = self.tacho_combobox.get().split(" - ")[0]
        except IndexError:
            messagebox.showwarning("Ошибка", "Выберите тахограф из списка")
            return
            
        calibration_date = self.calibration_date.get_date()
        next_calibration = self.next_calibration.get_date()
        
        if not all([tacho_id, calibration_date, next_calibration]):
            messagebox.showwarning("Ошибка", "Заполните все поля")
            return
        
        try:
            self.parent.database.add_record('calibration', {
                'tachograph_id': tacho_id,
                'calibration_date': calibration_date,
                'next_calibration_date': next_calibration,
                'user_id': self.user_id,
                'completed': False
            })
            self.parent.load_calibration_data()
            self.destroy()
            messagebox.showinfo("Успех", "Калибровка добавлена")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

class MainView:
    def __init__(self, root: Tk.Tk, user_id, account_type):
        self.root = root
        self.user_id = user_id
        self.account_type = account_type
        self.root.title(f'Тахограф 13 - {account_type.capitalize()}')
  
        self.root.minsize(800, 600)  
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        self.database = Database()
        self.notebook = ttk.Notebook(self.root)
        self.create_notifications_tab()
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        exit_frame = ttk.Frame(self.root)
        exit_frame.pack(side=Tk.TOP, fill=Tk.X, padx=5, pady=5)
        ttk.Button(exit_frame, text="Выйти", command=self.logout).pack(side=Tk.RIGHT)

        self.activation_data = []
        self.repair_data = []
        self.calibration_data = []
        
        self.permissions = {
            'admin': {
                'tables': ['users', 'contact', 'mrp', 'passport', 'tachograph', 'vehicle'],
                'tabs': ['activation', 'repair', 'calibration']
            },
            'operator': {
                'tables': ['contact', 'mrp', 'passport', 'tachograph', 'vehicle'],
                'tabs': ['activation']
            },
            'master': {
                'tables': ['contact', 'tachograph', 'vehicle'],
                'tabs': ['repair', 'calibration', 'activation']
            }
        }
        
        self.column_translations = {
            'users': {
                'id': 'ID',
                'username': 'Логин',
                'password': 'Пароль',
                'account_type': 'Тип учётки'
            },
            'contact': {
                'id': 'ID',
                'full_name': 'ФИО',
                'SNILS': 'СНИЛС',
                'TIN': 'ИНН',
                'phone': 'Телефон',
                'PTS': 'ПТС',
                'STS': 'СТС',
                'Client_id': 'ID клиента'
            },
            'mrp': {
                'id': 'ID',
                'source_path': 'Номер доверености',
                'date_start': 'Дата начала',
                'date_end': 'Дата окончания',
                'Client_id': 'ID клиента',
                'contact_id': 'ID контакта'

            },
            'passport': {
                'id': 'ID',
                'series_number': 'Серия-номер',
                'issued': 'Кем выдан',
                'dpt_code': 'Код подразделения',
                'date_issued': 'Дата выдачи',
                'full_name': 'ФИО',
                'gender': 'Пол',
                'date_of_birth': 'Дата рождения',
                'place_of_birth': 'Место рождения',
                'Contact_id': 'ID контакта',
                'last_used': 'Последнее использование'
            },
            'tachograph': {
                'id': 'ID',
                'manufacturer': 'Производитель',
                'model': 'Модель',
                'serial_number': 'Серийный номер',
                'vehicle_id': 'ID транспорта',
                'contact_id': 'ID контакта'
            },
            'vehicle': {
                'id': 'ID',
                'brand': 'Марка',
                'model': 'Модель',
                'gos_number': 'Гос. номер',
                'PTS': 'ПТС',
                'STS': 'СТС'
            }
        }
        
        
        allowed_tabs = self.permissions[self.account_type]['tabs']
        if 'activation' in allowed_tabs:
            self.create_activation_tab()
        if 'repair' in allowed_tabs:
            self.create_repair_tab()
        if 'calibration' in allowed_tabs:
            self.create_calibration_tab()
        
        allowed_tables = self.permissions[self.account_type]['tables']
        if allowed_tables:
            self.main_frame = ttk.Frame(self.notebook)
            self.notebook.add(self.main_frame, text="Основные таблицы")
            self.init_main_tab_ui(allowed_tables)
        
        self.notebook.pack(expand=1, fill='both')

    def logout(self):

        if hasattr(self, 'cleanup_thread'):
            self.cleanup_thread.stop()
            self.cleanup_thread.join(timeout=2.0)

        if hasattr(self, 'database'):
            self.database.close()

        for widget in self.root.winfo_children():
            widget.destroy()
        
        login = LoginWindow(self.root)
        self.root.wait_window(login)
        
        if login.success:
            MainView(self.root, login.user_id, login.account_type)
        else:
            self.root.destroy()

    def on_tab_changed(self, event):
        if self.notebook.tab(self.notebook.select(), "text") == "Уведомления":
            self.load_notifications()

    def create_notifications_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Уведомления")
        
        columns = ('Тип', 'Сообщение', 'Дата')
        self.notifications_tree = ttk.Treeview(frame, columns=columns, show='headings')
        for col in columns:
            self.notifications_tree.heading(col, text=col)
            self.notifications_tree.column(col, width=200)
        
        self.notifications_tree.pack(fill='both', expand=True)
        self.load_notifications()

    def load_notifications(self):
        self.notifications_tree.delete(*self.notifications_tree.get_children())
        try:
            notifications = self.database.get_notifications()
            today = datetime.now().date()
            
            for note in notifications:
                tags = []
                end_date = note['next_calibration_date']
                
                if isinstance(end_date, datetime):
                    end_date = end_date.date()
                
                delta = (end_date - today).days
                note_type = 'Калибровка через 2 недели'
                tags.append('calibration')
                message = f"Тахограф {note['serial_number']} (ID: {note['id']})"
                
                end_date_str = end_date.strftime('%d.%m.%Y')
                
                self.notifications_tree.insert("", Tk.END, 
                    values=(note_type, message, end_date_str),
                    tags=tuple(tags))
                    
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка загрузки уведомлений: {str(e)}")

    def update_tree(self):
        self.tree.delete(*self.tree.get_children())
        for record in self.database.get_table_data(self.current_table):
            formatted_values = []
            tags = []
            for col, value in record.items():
                if self.current_table == 'mrp' and col == 'date_end':
                    end_date = value.date() if isinstance(value, datetime) else value
                    today = datetime.now().date()
                    delta = (end_date - today).days if end_date >= today else 0

                    if end_date < today:
                        tags.append('expired')
                    elif delta <= 14:
                        tags.append('two_weeks')

                if col == 'legal_entity':
                    formatted_values.append('Юр.лицо' if value == b'\x01' else 'Физ.лицо')
                elif isinstance(value, (datetime, date)):
                    formatted_values.append(value.strftime('%Y-%m-%d'))
                else:
                    formatted_values.append(value)

        self.tree.insert("", Tk.END, values=tuple(formatted_values), tags=tuple(tags))

    def search_records(self):
        query = self.search_entry.get().lower()
        self.tree.delete(*self.tree.get_children())
        for record in self.current_data:
            if any(query in str(value).lower() for value in record.values()):
                formatted_values = []
                for col, value in record.items():
                    if col == 'legal_entity':
                        formatted_values.append('Юр.лицо' if value == b'\x01' else 'Физ.лицо')
                    elif isinstance(value, (datetime, date)):
                        formatted_values.append(value.strftime('%Y-%m-%d'))
                    else:
                        formatted_values.append(value)
                self.tree.insert("", Tk.END, values=tuple(formatted_values))

    def init_main_tab_ui(self, allowed_tables):
        buttons_frame = Tk.Frame(self.main_frame)
        buttons_frame.pack(fill='x', padx=5, pady=5)

        buttons = [
            ('Пользователи', 'users'), 
            ('Контакты', 'contact'), 
            ('МЧД', 'mrp'),
            ('Паспорта', 'passport'), 
            ('Тахографы', 'tachograph'),
            ('Транспорт', 'vehicle'),
        ]
        
        filtered_buttons = [(text, table) for text, table in buttons if table in allowed_tables]
        
        for idx, (text, table) in enumerate(filtered_buttons):
            ttk.Button(buttons_frame, text=text, 
                     command=lambda t=table: self.change_table(t)).grid(row=0, column=idx, padx=2)

        search_frame = ttk.Frame(buttons_frame)
        search_frame.grid(row=1, column=0, columnspan=len(filtered_buttons), sticky='ew', pady=5)
        
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side='left', padx=5, fill='x', expand=True)
        
        ttk.Button(search_frame, text="Поиск", 
                 command=self.search_records).pack(side='left', padx=5)
        ttk.Button(search_frame, text="Показать все", 
                 command=self.update_tree).pack(side='left', padx=5)

        self.tree_frame = Tk.Frame(self.main_frame)
        self.tree_frame.pack(fill='both', expand=True)
        self.tree = ttk.Treeview(self.tree_frame)
        self.tree.pack(fill='both', expand=True)
        
        if self.account_type in ['admin', 'operator', 'master']:
            control_frame = Tk.Frame(self.main_frame)
            control_frame.pack(fill='x', pady=5)
            ttk.Button(control_frame, text="Добавить", command=self.to_add_view).grid(row=0, column=0, padx=2)
            ttk.Button(control_frame, text='Изменить', command=self.to_edit_view).grid(row=0, column=1, padx=2)
            ttk.Button(control_frame, text='Удалить', command=self.to_delete_view).grid(row=0, column=2, padx=2)
        
        if filtered_buttons:
            self.change_table(filtered_buttons[0][1])

    def create_activation_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Активация")

        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side='left', padx=5)
        
        self.activation_search_entry = ttk.Entry(search_frame, width=30)
        self.activation_search_entry.pack(side='left', padx=5)
        
        ttk.Button(search_frame, text="Поиск", 
                 command=lambda: self.search_activation()).pack(side='left', padx=5)
        ttk.Button(search_frame, text="Показать все", 
                 command=self.load_activation_data).pack(side='left', padx=5)
        
        self.activation_tree = ttk.Treeview(frame, columns=('ID', 'ФИО', 'Телефон', 'Дата', 'Статус'), show='headings')
        for col in ['ID', 'ФИО', 'Телефон', 'Дата', 'Статус']:
            self.activation_tree.heading(col, text=col)
        self.activation_tree.pack(fill='both', expand=True)
        
        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        if self.account_type in ['admin', 'operator', 'master']:
            ttk.Button(control_frame, text="Новая активация", command=self.add_activation).grid(row=0, column=0, padx=2)
            ttk.Button(control_frame, text="Удалить", command=lambda: self.delete_record('activation', self.activation_tree)).grid(row=0, column=1, padx=2)
            ttk.Button(control_frame, text="Отметить выполненной", command=lambda: self.mark_procedure_completed('activation')).grid(row=0, column=2, padx=2)
        
        self.load_activation_data()

    def search_activation(self):
        query = self.activation_search_entry.get().lower()
        self.activation_tree.delete(*self.activation_tree.get_children())
        for record in self.activation_data:
            if (query in str(record['id']).lower() or
                query in record['full_name'].lower() or
                query in record['phone'].lower() or
                query in record['activation_datetime'].strftime('%Y-%m-%d %H:%M').lower() or
                query in ("выполнено" if record['completed'] else "активно").lower()):
                self.activation_tree.insert("", Tk.END, values=(
                    record['id'], 
                    record['full_name'],
                    record['phone'], 
                    record['activation_datetime'].strftime('%Y-%m-%d %H:%M'),
                    "Выполнено" if record['completed'] else "Активно"
                ))

    def create_repair_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Ремонт")

        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side='left', padx=5)
        
        self.repair_search_entry = ttk.Entry(search_frame, width=30)
        self.repair_search_entry.pack(side='left', padx=5)
        
        ttk.Button(search_frame, text="Поиск", 
                 command=lambda: self.search_repair()).pack(side='left', padx=5)
        ttk.Button(search_frame, text="Показать все", 
                 command=self.load_repair_data).pack(side='left', padx=5)
        
        columns = ('ID', 'ФИО', 'Телефон', 'Серийный номер', 'Дата', 'Ответственный')
        self.repair_tree = ttk.Treeview(frame, columns=columns, show='headings')
        for col in columns: self.repair_tree.heading(col, text=col)
        self.repair_tree.pack(fill='both', expand=True)
        
        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        if self.account_type in ['admin', 'master']:
            ttk.Button(control_frame, text="Новый ремонт", command=self.add_repair).grid(row=0, column=0, padx=2)
            ttk.Button(control_frame, text="Удалить", command=lambda: self.delete_record('repair', self.repair_tree)).grid(row=0, column=1, padx=2)
            ttk.Button(control_frame, text="Экспорт акта", command=self.export_repair_act).grid(row=0, column=3, padx=2)
        
        self.load_repair_data()

    def search_repair(self):
        query = self.repair_search_entry.get().lower()
        self.repair_tree.delete(*self.repair_tree.get_children())
        for record in self.repair_data:
            if  query in record['full_name'].lower():
                self.repair_tree.insert("", Tk.END, values=(
                    record['id'], 
                    record['full_name'], 
                    record['phone'],
                    record['serial_number'], 
                    record['repair_datetime'].strftime('%Y-%m-%d %H:%M'),
                    record['username']
                ))


    def create_calibration_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Калибровка")

        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        
        search_frame = ttk.Frame(control_frame)
        search_frame.pack(side='left', padx=5)
        
        self.calibration_search_entry = ttk.Entry(search_frame, width=30)
        self.calibration_search_entry.pack(side='left', padx=5)
        
        ttk.Button(search_frame, text="Поиск", 
                 command=lambda: self.search_calibration()).pack(side='left', padx=5)
        ttk.Button(search_frame, text="Показать все", 
                 command=self.load_calibration_data).pack(side='left', padx=5)
        
        columns = ('ID', 'Серийный номер', 'Дата', 'Следующая', 'Ответственный', 'Статус')
        self.calibration_tree = ttk.Treeview(frame, columns=columns, show='headings')
        for col in columns: self.calibration_tree.heading(col, text=col)
        self.calibration_tree.pack(fill='both', expand=True)
        
        control_frame = Tk.Frame(frame)
        control_frame.pack(fill='x', pady=5)
        if self.account_type in ['admin', 'master']:
            ttk.Button(control_frame, text="Новая калибровка", command=self.add_calibration).grid(row=0, column=0, padx=2)
            ttk.Button(control_frame, text="Удалить", command=lambda: self.delete_record('calibration', self.calibration_tree)).grid(row=0, column=1, padx=2)
            ttk.Button(control_frame, text="Отметить выполненной", command=lambda: self.mark_procedure_completed('calibration')).grid(row=0, column=2, padx=2)
            ttk.Button(control_frame, text="Экспорт отчета", command=self.export_calibration_report).grid(row=0, column=3, padx=2)
        
        self.load_calibration_data()

    def search_calibration(self):
        query = self.calibration_search_entry.get().lower()
        self.calibration_tree.delete(*self.calibration_tree.get_children())
        for record in self.calibration_data:
            if query in record['serial_number'].lower() or query in ("выполнено" if record['completed'] else "активно").lower():
                self.calibration_tree.insert("", Tk.END, values=(
                    record['id'], 
                    record['serial_number'],
                    record['calibration_date'].strftime('%Y-%m-%d'),
                    record['next_calibration_date'].strftime('%Y-%m-%d'),
                    record['username'],
                    "Выполнено" if record['completed'] else "Активно"
                ))

    def delete_record(self, table_type, tree):
        selected_item = tree.selection()
        if not selected_item: return
        
        if messagebox.askyesno("Подтверждение", "Удалить запись?"):
            record_id = tree.item(selected_item[0])['values'][0]
            try:
                self.database.delete_record(table_type, record_id)
                if table_type == 'activation': 
                    self.load_activation_data()
                elif table_type == 'repair': 
                    self.load_repair_data()
                elif table_type == 'calibration': 
                    self.load_calibration_data()
                messagebox.showinfo("Успех", "Запись удалена")
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

    def add_activation(self): AddActivationWindow(self)
    def add_repair(self): AddRepairWindow(self, self.user_id)
    def add_calibration(self): AddCalibrationWindow(self, self.user_id)

    def mark_procedure_completed(self, procedure_type):
        tree = getattr(self, f'{procedure_type}_tree')
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Ошибка", "Выберите запись")
            return
        
        record_id = tree.item(selected_item[0])['values'][0]
        
        if self.database.mark_procedure_completed(procedure_type, record_id):
            messagebox.showinfo("Успех", "Процедура отмечена выполненной")
            
            if procedure_type == 'activation':
                self.load_activation_data()
                for record in self.activation_data:
                    if record['id'] == record_id:
                        self.database.delete_passport_by_contact(record['contact_id'])
                        break
            elif procedure_type == 'calibration':
                self.load_calibration_data()
                for record in self.calibration_data:
                    if record['id'] == record_id:
                        self.database.cursor.execute(
                            "SELECT contact_id FROM tachograph WHERE id IN "
                            "(SELECT tachograph_id FROM calibration WHERE id = %s)",
                            (record_id,)
                        )
                        result = self.database.cursor.fetchone()
                        if result:
                            self.database.delete_passport_by_contact(result['contact_id'])
                        break
    def load_activation_data(self):
        self.activation_data = self.database.get_table_data('activation')
        self.activation_tree.delete(*self.activation_tree.get_children())
        for record in self.activation_data:
            self.activation_tree.insert("", Tk.END, values=(
                record['id'], 
                record['full_name'],
                record['phone'], 
                record['activation_datetime'].strftime('%Y-%m-%d %H:%M'),
                "Выполнено" if record['completed'] else "Активно"
            ))

    def load_repair_data(self):
        self.repair_data = self.database.get_table_data('repair')
        self.repair_tree.delete(*self.repair_tree.get_children())
        for record in self.repair_data:
            self.repair_tree.insert("", Tk.END, values=(
                record['id'], 
                record['full_name'], 
                record['phone'],
                record['serial_number'], 
                record['repair_datetime'].strftime('%Y-%m-%d %H:%M'),
                record['username']
            ))

    def load_calibration_data(self):
        self.calibration_data = self.database.get_table_data('calibration')
        self.calibration_tree.delete(*self.calibration_tree.get_children())
        for record in self.calibration_data:
            self.calibration_tree.insert("", Tk.END, values=(
                record['id'], 
                record['serial_number'],
                record['calibration_date'].strftime('%Y-%m-%d'),
                record['next_calibration_date'].strftime('%Y-%m-%d'),
                record['username'],
                "Выполнено" if record['completed'] else "Активно"
            ))

    def change_table(self, table):
        self.current_table = table
        self.columns = self.database.get_table_columns(table)

        self.current_data = self.database.get_table_data(table)
        
        translated_columns = [self.column_translations[table].get(col, col) for col in self.columns]
        
        self.tree.destroy()
        self.tree = ttk.Treeview(self.tree_frame, columns=self.columns, show='headings')
        
        for col, translated in zip(self.columns, translated_columns):
            self.tree.heading(col, text=translated)
            self.tree.column(col, width=120, anchor='center')
        
        self.tree.pack(fill='both', expand=True)
        self.update_tree()

    def update_tree(self):
        self.tree.delete(*self.tree.get_children())

        self.current_data = self.database.get_table_data(self.current_table)
        
        for record in self.current_data:
            formatted_values = []
            for col, value in record.items():
                if col == 'legal_entity':
                    formatted_values.append('Юр.лицо' if value == b'\x01' else 'Физ.лицо')
                elif isinstance(value, (datetime, date)):
                    formatted_values.append(value.strftime('%Y-%m-%d'))
                else:
                    formatted_values.append(value)
            self.tree.insert("", Tk.END, values=tuple(formatted_values))

    def to_add_view(self):
        if self.account_type != 'admin':
            messagebox.showwarning("Доступ запрещен", "У вас недостаточно прав")
            return
        AddView(self, self.current_table, [c for c in self.columns if c != 'id'])

    def to_edit_view(self):
        if self.account_type != 'admin':
            messagebox.showwarning("Доступ запрещен", "У вас недостаточно прав")
            return
        if item := self.tree.focus():
            record_id = self.tree.item(item)['values'][0]
            if record := self.database.get_record(self.current_table, record_id):
                EditView(self, self.current_table, record)

    def to_delete_view(self):
        if self.account_type != 'admin':
            messagebox.showwarning("Доступ запрещен", "У вас недостаточно прав")
            return
        if item := self.tree.focus():
            record_id = self.tree.item(item)['values'][0]
            if self.database.delete_record(self.current_table, record_id):
                self.update_tree()

    def export_calibration_report(self):
        selected_item = self.calibration_tree.selection()
        if not selected_item:
            messagebox.showwarning("Ошибка", "Выберите запись для экспорта")
            return
        
        calibration_id = self.calibration_tree.item(selected_item[0])['values'][0]
        data = self.database.get_calibration_details(calibration_id)
        
        if not data:
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx")],
            title="Сохранить отчет о калибровке"
        )
        
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Калибровка"

            ws.merge_cells('A1:F1')
            ws['A1'] = "Отчет о проведении настройки тахографа"
            ws['A1'].font = Font(bold=True, size=14)
            
            owner_data = [
                ("Владелец ТС:", data['full_name']),
                ("Адрес:", "170006, Тверская обл., г.Тверь, ул.Достоевского д.13 корп.А оф.13"),
                ("ТС (марка/модель):", f"{data['brand']} {data['model']}"),
                ("Год выпуска ТС:", ""),
                ("Гос. Номер:", data['gos_number']),
                ("Показания одометра (км):", ""),
                ("Тахограф (модель):", "VDO 3283.421"),
                ("Зав. номер:", data['serial_number']),
                ("СКЗИ:", data['serial_number'])
            ]
            
            for row, (label, value) in enumerate(owner_data, start=3):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

            ws['A12'] = "Данные по настройке"
            calibration_data = [
                ("Типоразмер шин (mm):", "425/85R21"),
                ("Давление в шинах (bar):", 7),
                ("Температура (°C):", 16),
                ("Высота протектора (mm):", 11)
            ]
            
            for row, (label, value) in enumerate(calibration_data, start=13):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

            ws.merge_cells('A17:E17')
            ws['A17'] = "Параметры настройки"
            params_data = [
                ("L (mm):", 3800),
                ("W (imp/km):", 4133),
                ("K (imp/km):", 4133)
            ]
            
            for col, (label, value) in enumerate(params_data, start=1):
                ws.cell(row=18, column=col*2-1, value=label)
                ws.cell(row=18, column=col*2, value=value)

            ws['A20'] = f"Дата настройки: {data['calibration_date'].strftime('%d.%m.%Y')}"
            ws['A21'] = f"Дата следующей настройки: {data['next_calibration_date'].strftime('%d.%m.%Y')}"
            ws['A22'] = f"Мастер: {data['username']}"

            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("Успех", "Отчет о калибровке сохранен")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте: {str(e)}")

    def export_repair_act(self):
        selected_item = self.repair_tree.selection()
        if not selected_item:
            messagebox.showwarning("Ошибка", "Выберите запись для экспорта")
            return
        
        repair_id = self.repair_tree.item(selected_item[0])['values'][0]
        data = self.database.get_repair_details(repair_id)
        
        if not data:
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файлы", "*.xlsx")],
            title="Сохранить акт ремонта"
        )
        
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Акт ремонта"

            ws.merge_cells('A1:F1')
            ws['A1'] = "АКТ ПРИЕМА-ПЕРЕДАЧИ ТАХОГРАФА ИЗ РЕМОНТА"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A3'] = "ИСПОЛНИТЕЛЬ:"
            executor_data = [
                ("Наименование", "ООО 'Рога и Копыта'"),
                ("Адрес", "170006, Тверская обл., г.Тверь, ул.Достоевского д.13 корп.А оф.13"),
                ("Телефон", "+7(1234)567-890"),
                ("Email", "info@rogandcop.ru")
            ]
            
            for row, (label, value) in enumerate(executor_data, start=4):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

            ws['A8'] = "ЗАКАЗЧИК:"
            customer_data = [
                ("Наименование", data['full_name']),
                ("Адрес", "г. Москва, ул. Пушкина, д.564"),
                ("Марка ТС", data['brand']),
                ("Гос. номер", "")
            ]
            
            for row, (label, value) in enumerate(customer_data, start=9):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

            ws['A13'] = "ДАННЫЕ КОНТРОЛЬНОГО УСТРОЙСТВА"
            device_data = [
                ("Модель", "VDO 3283.421"),
                ("Серийный номер", data['serial_number']),
                ("Номер СКЗИ", "211350007977573"),
                ("Гарантия", "негарантийный")
            ]
            
            for row, (label, value) in enumerate(device_data, start=14):
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value

            ws['A18'] = "ВЫПОЛНЕННЫЕ РАБОТЫ"
            works = [
                "Диагностика",
                "Замена батарейки",
                "Обновление ПО"
            ]
            
            for row, work in enumerate(works, start=19):
                ws[f'A{row}'] = work

            ws['A22'] = "Исполнитель: _________________ ()"
            ws['A23'] = "Заказчик: ____________________"
            ws['A24'] = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"

            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("Успех", "Акт ремонта сохранен")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при экспорте: {str(e)}")


if __name__ == '__main__':
    root = Tk.Tk()
    root.withdraw()
    
    login = LoginWindow(root)
    root.wait_window(login)
    
    if login.success:
        main_view = MainView(root, login.user_id, login.account_type)
        root.deiconify()
        root.mainloop()
    else:
        root.destroy()